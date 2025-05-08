from django.shortcuts import redirect, render
from .models import Machine, PieceRechange, MesureCapteur, Notification, Intervention
from . import models
from django.db.models import Q
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.shortcuts import get_object_or_404
from datetime import date
from django.views.decorators.csrf import csrf_exempt
from .models import PieceRechange
from django.contrib import messages
from django.db.models import Count, Avg, Sum
from django.utils.timezone import now
from collections import defaultdict
import calendar
from .forms import MachineForm
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from datetime import date, timedelta
from .models import PlanMaintenance
from .forms import PlanMaintenanceForm
from datetime import datetime, timedelta
from django.contrib.auth.models import User
from .models import PlanMaintenance, Machine
from django.contrib import messages



def accueil(request):
    today = date.today()
    dans_3_jours = today + timedelta(days=3)

    prochaines_interventions = Intervention.objects.filter(
        date_planifiee__range=(today, dans_3_jours)
    )

    interventions = Intervention.objects.all()
    total_minutes = sum(i.duree_total for i in interventions if i.duree_total)
    total_heures = total_minutes // 60
    minutes_restant = total_minutes % 60

    mttr = round(total_minutes / len(interventions)) if interventions else 0
    mttr_heures = mttr // 60
    mttr_minutes = mttr % 60

    return render(request, 'accueil.html', {
        'prochaines_interventions': prochaines_interventions,
        'total_interventions': interventions.count(),
        'total_heures': total_heures,
        'total_minutes': minutes_restant,
        'mttr_heures': mttr_heures,
        'mttr_minutes': mttr_minutes
    })

def machines(request):
    machines = Machine.objects.all()

    nom = request.GET.get('nom')
    statut = request.GET.get('statut')
    type_machine = request.GET.get('type')
    emplacement = request.GET.get('emplacement')

    if nom:
        machines = machines.filter(nom__icontains=nom)
    if statut:
        machines = machines.filter(statut=statut)
    if type_machine:
        machines = machines.filter(type__icontains=type_machine)
    if emplacement:
        machines = machines.filter(emplacement__icontains=emplacement)

    return render(request, 'machines.html', {'machines': machines})

def ajouter_machine(request):
    if request.method == 'POST':
        nom = request.POST.get('nom')
        type_machine = request.POST.get('type')
        statut = request.POST.get('statut')
        derniere_maintenance = request.POST.get('derniere_maintenance')
        frequence_jours = request.POST.get('frequence_jours')
        emplacement = request.POST.get('emplacement')

        Machine.objects.create(
            nom=nom,
            type=type_machine,
            statut=statut,
            derniere_maintenance=derniere_maintenance,
            frequence_jours=frequence_jours,
            emplacement=emplacement,
        )
        messages.success(request, "✅ Machine ajoutée avec succès !")
        return redirect('machines')

    return render(request, 'ajouter_machine.html')


def pieces(request):
    query = request.GET.get('q', '')
    pieces = PieceRechange.objects.filter(
        Q(code__icontains=query) | Q(designation__icontains=query)
    )
    return render(request, 'pieces.html', {'pieces': pieces, 'query': query})
def capteurs(request):
    capteurs = MesureCapteur.objects.all()
    return render(request, 'capteurs.html', {'capteurs': capteurs})

def interventions(request):
    interventions = Intervention.objects.all()
    return render(request, 'interventions.html', {'interventions': interventions})

def notifications(request):
    notifications = Notification.objects.all()
    return render(request, 'notifications.html', {'notifications': notifications})

def import_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = load_workbook(file)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not row or len(row) < 6:
                continue  # saute les lignes incomplètes

            code, designation, quantite, seuil_alerte, prix_unitaire, machines_str = row

            # Gestion du prix_unitaire avec Decimal
            try:
                prix_unitaire = Decimal(str(prix_unitaire)).quantize(Decimal('0.01')) if prix_unitaire else Decimal('0.00')
            except InvalidOperation:
                prix_unitaire = Decimal('0.00')

            # Création ou récupération de la pièce
            piece, created = PieceRechange.objects.get_or_create(
                code=code,
                defaults={
                    'designation': designation,
                    'quantite_stock': quantite or 0,
                    'seuil_alerte': seuil_alerte or 0,
                    'prix_unitaire': prix_unitaire
                }
            )

            if not created:
                continue  # Ignore les doublons (déjà existants)

            # Ajout des machines compatibles
            if machines_str:
                noms_machines = machines_str.split(',')
                for nom_machine in noms_machines:
                    nom_machine = nom_machine.strip()
                    try:
                        machine = Machine.objects.get(nom=nom_machine)
                        piece.machines.add(machine)
                    except Machine.DoesNotExist:
                        print(f"⚠️ Machine '{nom_machine}' introuvable.")

        return redirect('pieces')  # Redirection vers la liste des pièces

    return render(request, 'import_excel.html')

def delete_piece(request, piece_id):
    piece = get_object_or_404(PieceRechange, id=piece_id)
    piece.delete()
    return redirect('pieces')

def delete_all_pieces(request):
    PieceRechange.objects.all().delete()
    return redirect('pieces')

def modifier_piece(request):
    if request.method == 'POST':
        piece_id = request.POST.get('piece_id')
        code = request.POST.get('code')
        designation = request.POST.get('designation')
        quantite = request.POST.get('quantite')
        seuil_alerte = request.POST.get('seuil')
        prix_unitaire = request.POST.get('prix')
        machines = piece.machines_compatibles.all()

        try:
            piece = PieceRechange.objects.get(id=piece_id)
            piece.code = code
            piece.designation = designation
            piece.quantite_stock = quantite
            piece.seuil_alerte = seuil_alerte
            piece.prix_unitaire = prix_unitaire
            piece.save()
            messages.success(request, "✅ Pièce modifiée avec succès !")
            if machines:
                piece.machines.set(machines)  # ✅ met à jour les machines associées

            piece.save()
            messages.success(request, "✅ Pièce modifiée avec succès !")
            
        except PieceRechange.DoesNotExist:
            pass  # ou log

    return render(request, 'pieces.html', {'piece': piece, 'machines_compatibles': machines})




def dashboard_maintenance(request):
    # Regrouper les maintenances par mois
    maintenances = PlanMaintenance.objects.all()
    data = defaultdict(lambda: {
        'count': 0,
        'duration_total': 0,
        'cost_total': 0,
        'on_time': 0
    })

    for m in maintenances:
        mois = (m.date_derniere_realisation or m.prochaine_date).strftime('%B')  # exemple : Avril
        data[mois]['count'] += 1
        data[mois]['duration_total'] += m.duree
        data[mois]['cost_total'] += m.cout
        if not m.on_time:  # suppose un booléen
            data[mois]['on_time'] += 1

    # Calculer les indicateurs
    stats = []
    for mois, d in data.items():
        stats.append({
            'mois': mois,
            'maintenances': d['count'],
            'temps_total': d['duration_total'],
            'cout_moyen': round(d['cost_total'] / d['count'], 2) if d['count'] else 0,
            'taux_conformite': round(d['on_time'] / d['count'], 2) if d['count'] else 0,
            'mtbf': round((30 * 24) / d['count'], 2),
            'mttr': round(d['duration_total'] / d['count'], 2)
        })

    stats = sorted(stats, key=lambda x: list(calendar.month_name).index(x['mois']))

    return render(request, 'dashboard.html', {'stats': stats})

def voir_maintenance_machine(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    maintenances = machine.maintenances.all()
    return render(request, 'voir_maintenance_machine.html', {
        'machine': machine,
        'maintenances': maintenances
    })
def modifier_machine(request, id):
    machine = get_object_or_404(Machine, id=id)
    if request.method == 'POST':
        form = MachineForm(request.POST, instance=machine)
        if form.is_valid():
            form.save()
            return redirect('machines')
    else:
        form = MachineForm(instance=machine)
    
    # Ajoute les maintenances liées ici
    maintenances = machine.maintenances.all()  # ou machine.maintenances_preventives.all() selon ton modèle

    return render(request, 'modifier_machine.html', {
        'form': form,
        'machine': machine,
        'maintenances': maintenances,
    })

def supprimer_machine(request, id):
    machine = get_object_or_404(Machine, id=id)
    machine.delete()
    return redirect('machines')

def details_machine(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    maintenances = PlanMaintenance.objects.filter(machine=machine)
    pieces = machine.piecerechange_set.all()
    return render(request, 'machines/details_machine.html', {
        'machine': machine,
        'maintenances': maintenances,
        'pieces': pieces
    })
def custom_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('accueil')  # redirection après connexion
        else:
            return render(request, 'login.html', {'error': True})
    return render(request, 'login.html')
def interventions(request):
    if request.method == "POST":
        machine_id = request.POST.get('machine')
        description = request.POST.get('description')
        date_planifiee = request.POST.get('date_planifiee')

        machine = Machine.objects.get(id=machine_id)

        Intervention.objects.create(
            machine=machine,
            technicien=request.user,
            description=description,
            date_planifiee=date_planifiee,
            statut='planifiee',
        )
        return redirect('interventions')

    # 🧠 Mise à jour automatique des statuts
    interventions = Intervention.objects.all()
    for i in interventions:
        if i.date_planifiee and i.date_planifiee <= date.today() and not i.terminee:
            i.statut = 'en_cours'
            i.save()

    machines = Machine.objects.all()
    interventions = interventions.order_by('-date_planifiee')

    return render(request, 'interventions.html', {
        'interventions': interventions,
        'machines': machines
    })

def cloturer_intervention(request, intervention_id):
    i = get_object_or_404(Intervention, id=intervention_id)
    i.terminee = True
    i.date_intervention = date.today()
    i.statut = 'terminee'
    i.save()
    return redirect('interventions')

def calendrier(request):
    return render(request, 'calendrier.html')

def modifier_intervention(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)

    if request.method == 'POST':
        intervention.description = request.POST.get('description')
        intervention.date_planifiee = request.POST.get('date_planifiee')
        intervention.save()
        return redirect('interventions')

    return render(request, 'modifier_intervention.html', {'intervention': intervention})


def supprimer_intervention(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)
    intervention.delete()
    return redirect('interventions')

def alertes_du_jour(request):
    today = date.today()
    upcoming = today + timedelta(days=3)
    interventions = Intervention.objects.filter(date_planifiee__range=(today, upcoming), statut='planifiee')

    data = [{
        'machine': i.machine.nom,
        'description': i.description or "Pas de description",
        'date': i.date_planifiee.strftime('%d %b %Y'),
    } for i in interventions]

    return JsonResponse({'interventions': data})


def api_interventions(request):
    interventions = Intervention.objects.all()
    data = [
        {
            "title": f"{i.machine.nom} - {i.description}",
            "start": i.date_planifiee.isoformat() if i.date_planifiee else None,
            "end": i.date_planifiee.isoformat() if i.date_planifiee else None,
        }
        for i in interventions if i.date_planifiee
    ]
    return JsonResponse(data, safe=False)


def creer_plan_maintenance(request):
    if request.method == "POST":
        equipement_id = request.POST.get('equipement')
        description = request.POST.get('description')
        reglementaire = bool(request.POST.get('reglementaire'))
        assignes_ids = request.POST.getlist('assignes')
        labels = request.POST.get('labels', '')

        temps_maintenance_h = int(request.POST.get('temps_maintenance_h', 0))
        temps_maintenance_m = int(request.POST.get('temps_maintenance_m', 0))
        temps_arret_h = int(request.POST.get('temps_arret_h', 0))
        temps_arret_m = int(request.POST.get('temps_arret_m', 0))

        checklist = request.POST.get('checklist', '')
        frequence = int(request.POST.get('frequence', 1))
        unite = request.POST.get('unite')
        date_debut = request.POST.get('date_debut')
        heure_debut = request.POST.get('heure_debut')

        # Check if mandatory fields exist
        if not equipement_id or not date_debut :
            messages.error(request, "🚨 Champ obligatoire manquant.")
            return redirect('creer_plan_maintenance')

        # Récupérer la machine
        machine = get_object_or_404(Machine, id=equipement_id)

        # Créer le plan
        plan = PlanMaintenance.objects.create(
            titre=f"Plan pour {machine.nom}",
            description=description,
            machine=machine,
            reglementaire=reglementaire,
            labels=labels,
            checklist=checklist,
            temps_maintenance_minutes=temps_maintenance_h * 60 + temps_maintenance_m,
            temps_arret_minutes=temps_arret_h * 60 + temps_arret_m,
            frequence=frequence,
            unite=unite,
            date_debut=date_debut,
            heure_debut=heure_debut,
        )

        # Lier les techniciens
        techniciens = User.objects.filter(id__in=assignes_ids)
        plan.assignes.set(techniciens)

        messages.success(request, "✅ Plan de maintenance créé avec succès !")
        return redirect('liste_plans_maintenance')

    else:
        machines = Machine.objects.all()
        techniciens = User.objects.all()
        return render(request, 'plans/creer_plan.html', {'machines': machines, 'techniciens': techniciens})
def liste_plans_maintenance(request):
    plans = PlanMaintenance.objects.all()
    for plan in plans:
        if not hasattr(plan, 'date_creation'):
            setattr(plan, 'date_creation', datetime.now())
    return render(request, 'liste_plans_maintenance.html', {'plans': plans})

def accueil(request):
    return render(request, 'accueil.html')

def success_plan(request):
    return render(request, 'plans/success_plan.html')

def voir_plan(request, id):
    plan = get_object_or_404(PlanMaintenance, id=id)
    return render(request, 'plans/voir_plan.html', {'plan': plan})

def modifier_plan(request, id):
    plan = get_object_or_404(PlanMaintenance, id=id)
    if request.method == 'POST':
        form = PlanMaintenanceForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            return redirect('liste_plans_maintenance')
    else:
        form = PlanMaintenanceForm(instance=plan)
    return render(request, 'core/modifier_plan.html', {'form': form})

def supprimer_plan(request, id):
    plan = get_object_or_404(PlanMaintenance, id=id)
    if request.method == 'POST':
        plan.delete()
        return redirect('liste_plans_maintenance')
    return render(request, 'core/supprimer_plan.html', {'plan': plan})